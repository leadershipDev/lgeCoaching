import sys, os
sys.stdout.reconfigure(encoding='utf-8')
os.chdir(r'd:\project\CoachingLanding')
import openpyxl
fname = [f for f in os.listdir('.') if f.endswith('.xlsx')][0]
print('file:', fname)
wb = openpyxl.load_workbook(fname, data_only=True, read_only=True)
print('sheets:', wb.sheetnames)
ws = wb['선배경영자']
for row in ws.iter_rows(min_row=4, max_row=30, values_only=True):
    if row[1]:
        print(row[1], '|', row[10], '|', row[11], '|', row[12])
wb.close()
