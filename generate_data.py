# -*- coding: utf-8 -*-
"""
generate_data.py — 엑셀 → data.js 변환
"""
import json, re, sys, os
from collections import Counter

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "(all)`26년_LG전자_사외코치Pool_260325.xlsx")
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "data.js")

FIRM_GROUP_MAP = {'개별':'개별','CiT':'CiT','한스':'한스코칭','자경원':'코칭경영원'}

def clean(val):
    if val is None: return None
    s = str(val).strip()
    if s in ('','na','NA','N/A','-',' -','- ','–','—','None'): return None
    if s.startswith('=VLOOKUP') or s.startswith('='): return None
    return s

def clean_bool(val):
    if val is None: return None
    s = str(val).strip()
    if s in ('O','○','o'): return 'O'
    if s in ('△','▲'): return '△'
    return None

def to_price(val):
    """숫자→정수, 텍스트('별도협의' 등)→문자열, 0/None→None"""
    if val is None: return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v > 0 else None
    s = str(val).strip()
    if s in ('','0','-'): return None
    try: return int(float(s)) if int(float(s)) > 0 else None
    except: return s if s else None

# ── 가격 추출 ──────────────────────────────────────────

def _raw_to_amt(raw):
    """'-' 문자열은 None으로, 나머지는 to_price로 변환"""
    if isinstance(raw, str) and raw.strip() in ('-', ''): return None
    return to_price(raw)

def _indi_amt(raw):
    """개별 코치용: 0이거나 비어있으면 '해당없음', '-'는 None(표시 안 함)"""
    if raw is None: return '해당없음'
    if isinstance(raw, (int, float)):
        v = int(raw)
        return '해당없음' if v == 0 else v
    s = str(raw).strip()
    if s in ('-', ''): return None
    if s == '0': return '해당없음'
    try:
        v = int(float(s))
        return '해당없음' if v == 0 else v
    except:
        return s or '해당없음'

def _clean_sub(s):
    """sub 셀의 줄바꿈 정리: 'Group1\n(C레벨)' → 'Group1 (C레벨)'"""
    if s is None: return None
    return ' '.join(s.split())

def _clean_note(val):
    """비고 텍스트 정리: 줄바꿈 보존, 각 줄 앞뒤 공백 제거"""
    if val is None: return None
    s = str(val).strip()
    if not s: return None
    lines = [l.strip() for l in s.splitlines()]
    lines = [l for l in lines if l]
    return '\n'.join(lines)

def resolve_merges(ws):
    """세로 방향(행 확장) 병합 셀만 값을 전파한 행 튜플 리스트 반환.
    가로 병합(같은 행, 열 확장)은 전파하지 않아 category/type 오염을 방지."""
    merge_map = {}
    for mrange in ws.merged_cells.ranges:
        if mrange.max_row <= mrange.min_row:
            continue  # 같은 행 병합(가로)은 무시
        top_val = ws.cell(mrange.min_row, mrange.min_col).value
        for r in range(mrange.min_row + 1, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merge_map[(r, c)] = top_val
    rows = []
    for row_cells in ws.iter_rows():
        vals = []
        for cell in row_cells:
            coord = (cell.row, cell.column)
            if coord in merge_map:
                vals.append(merge_map[coord])
            else:
                vals.append(getattr(cell, 'value', None))
        rows.append(tuple(vals))
    return rows

def prices_cit(ws):
    """CiT: col[0]=category, col[1]=type, col[2]=amount, col[3]=note"""
    rows = resolve_merges(ws)
    result, category = [], None
    data = False
    for row in rows:
        if not data:
            if str(row[0] or '').strip() == '항목': data = True
            continue
        if str(row[0] or '').startswith('■'): break
        if not any(v is not None for v in row): continue
        cat = clean(row[0])
        typ = clean(row[1])
        amt = _raw_to_amt(row[2])
        note = _clean_note(row[3]) if len(row) > 3 else None
        if cat: category = cat
        if not category: continue
        if not cat and typ is None: continue
        result.append({"category": category, "sub": None, "type": typ, "amount": amt, "note": note})
    return result

def prices_hans(ws):
    """한스코칭: col[0]=category, col[1]=type, col[2]=amount, col[4+]=coach cols
    'X' 표시된 항목은 amount=None(미제공)으로 포함"""
    rows = resolve_merges(ws)
    header_idx = next((i for i,r in enumerate(rows) if str(r[0] or '').strip()=='항목'), None)
    if header_idx is None: return {}
    h = rows[header_idx]
    # col[4]부터 코치명, ' 코치' 접미어 제거
    coach_cols = [(j, str(h[j]).strip().replace(' 코치', ''))
                  for j in range(4, len(h)) if clean(h[j])]
    base, category = [], None
    for row in rows[header_idx+1:]:
        if str(row[0] or '').startswith('■'): break
        if not any(v is not None for v in row): continue
        cat = clean(row[0])
        typ = clean(row[1])
        amt = _raw_to_amt(row[2])
        note = _clean_note(row[3]) if len(row) > 3 else None
        if cat: category = cat
        if not category: continue
        if not cat and typ is None: continue
        base.append({"category": category, "type": typ, "amount": amt, "note": note, "_row": row})
    prices_by_name = {}
    for col_idx, name in coach_cols:
        coach_prices = []
        for bp in base:
            row = bp["_row"]
            cell = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ''
            excluded = cell in ('X', 'x', '미제공')
            coach_prices.append({
                "category": bp["category"], "sub": None,
                "type": bp["type"],
                "amount": '해당없음' if excluded else bp["amount"],
                "note": bp["note"]
            })
        prices_by_name[name] = coach_prices
    return prices_by_name

def prices_jakyungwon(ws):
    """자경원: col[0]=category, col[1]=sub, col[2]=type, col[3]=amount (sub 필드 보존)"""
    rows = resolve_merges(ws)
    result, category, sub = [], None, None
    data = False
    for row in rows:
        if not data:
            if str(row[0] or '').strip() == '항목': data = True
            continue
        if str(row[0] or '').startswith('■'): break
        cat = clean(row[0])
        s   = clean(row[1])
        typ = clean(row[2])
        amt = _raw_to_amt(row[3])
        note = _clean_note(row[4]) if len(row) > 4 else None
        if cat: category = cat; sub = None
        if s:   sub = _clean_sub(s)
        if not category: continue
        if not cat and s is None and typ is None: continue
        result.append({"category": category, "sub": sub, "type": typ, "amount": amt, "note": note})
    return result

def prices_individual(ws):
    """개별: header row에 코치명 배열, 각 행에 코치별 단가 (0→해당없음, note 포함)"""
    rows = resolve_merges(ws)
    header_idx = next((i for i,r in enumerate(rows) if str(r[0] or '').strip()=='항목'), None)
    if header_idx is None: return {}
    h = rows[header_idx]
    coach_cols = []
    note_col = None
    for j in range(2, len(h)):
        name = clean(h[j])
        if name and '비고' in name:
            note_col = j  # 마지막 비고1 열
        elif name:
            coach_cols.append((j, name))
    prices_by_name = {name: [] for _, name in coach_cols}
    category = None
    for row in rows[header_idx+1:]:
        cat = clean(row[0])
        typ = clean(row[1])
        if cat:
            if cat.startswith('비고'): break
            category = cat
        if category is None: continue
        note = _clean_note(row[note_col]) if note_col and note_col < len(row) else None
        for col_idx, name in coach_cols:
            raw = row[col_idx] if col_idx < len(row) else None
            amt = _indi_amt(raw)
            prices_by_name[name].append({
                "category": category, "sub": None,
                "type": typ, "amount": amt, "note": note
            })
    return prices_by_name

def coaches_sunbae(ws):
    """선배경영자 시트 파싱
    col[0]=No, col[1]=성명, col[2]=alias,
    col[3]=직위, col[4]=회사, col[5]=부서, col[6]=직책, col[7]=퇴임일자,
    col[8]=주요경력, col[9]=사업가, col[10]=FL, col[11]=리더십,
    col[12]=코치자격, col[13]=비고, col[14]=풀구축년도
    데이터는 4행부터 시작
    """
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        no = row[0]
        if no is None: continue
        try: no = int(no)
        except: continue
        name      = clean(row[1]) or ''
        alias     = clean(row[2]) or ''
        title     = clean(row[3])
        company   = clean(row[4])
        dept      = clean(row[5])
        position  = clean(row[6])
        retire_dt = clean(row[7])
        background= clean(row[8])
        area_biz  = clean(row[9])
        area_fl   = clean(row[10])
        area_lead = clean(row[11])
        cert      = clean(row[12])
        note      = _clean_note(row[13])
        pool_year = clean(row[14]) if len(row) > 14 else None
        result.append({
            "no": no, "firm": "선배경영자", "group": "선배경영자",
            "name": name, "alias": alias,
            "birthYear": None, "gender": None,
            "hasPhoto": False,
            "profileUrl": None, "introVideoUrl": None, "youtubeUrl": None,
            "contact": None, "email": None,
            "title": title, "company": company, "dept": dept, "position": position,
            "retireDate": retire_dt, "poolYear": pool_year,
            "recentLgActivity": None, "background": background,
            "strengths": None, "lgExperience": None,
            "successCases": None, "coachingDomain": None,
            "visionStrategy": None, "performanceExec": None,
            "talentOrg": None, "selfDev": None,
            "cLevel": None, "execCoaching": None, "midManager": None,
            "diagnosticTools": None,
            "groupCoaching": None, "regionalTravel": None, "languages": None,
            "lgElecExperience": None,
            "certKca": cert, "certDomesticOther": None, "certIcf": None,
            "areaBusiness": area_biz, "areaFL": area_fl, "areaLeadership": area_lead,
            "note": note,
            "prices": [],
        })
    return result

# ── 메인 ──────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_names = wb.sheetnames
    def ws(name): return wb[name] if name in sheet_names else None

    ws_main = wb.worksheets[0]
    p_cit  = prices_cit(wb['비용기준(CiT)'])           if '비용기준(CiT)' in sheet_names else []
    p_hans = prices_hans(wb['비용기준(한스)'])           if '비용기준(한스)' in sheet_names else {}
    p_jaky = prices_jakyungwon(wb['비용기준(코경원)'])   if '비용기준(코경원)' in sheet_names else []
    p_indi = prices_individual(wb['비용기준(개별)'])     if '비용기준(개별)' in sheet_names else {}
    wb_ro  = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sunbae = coaches_sunbae(wb_ro['선배경영자'])          if '선배경영자' in wb_ro.sheetnames else []
    wb_ro.close()

    def cell_val(row, idx):
        return row[idx].value if idx < len(row) else None

    def cell_hyperlink(row, idx):
        """셀 값이 URL이 아닌 텍스트(예: 'link')인 경우 hyperlink target을 반환"""
        if idx >= len(row): return None
        cell = row[idx]
        val = clean(cell.value)
        if val and val.startswith('http'): return val
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target
        return val

    coaches = []
    for row in ws_main.iter_rows(min_row=3):
        no = cell_val(row, 0)
        if no is None: continue
        try: no = int(no)
        except: continue

        firm_raw = clean(cell_val(row, 1)) or ''
        firm  = firm_raw if firm_raw else '개별'
        group = FIRM_GROUP_MAP.get(firm, firm)
        name  = clean(cell_val(row, 2)) or ''
        alias = clean(cell_val(row, 3)) or ''

        by = cell_val(row, 4)
        if by:
            m = re.search(r'\d{4}', str(by))
            by = int(m.group()) if m else None
        else: by = None

        gender_raw = clean(cell_val(row, 5))
        # gender 열에 연도가 잘못 들어온 경우 → birthYear로 이동
        if gender_raw and re.fullmatch(r'\d{4}', gender_raw):
            if by is None:
                by = int(gender_raw)
            gender_raw = None
        gender = gender_raw
        photo_file   = clean(cell_val(row, 6))
        intro_link   = cell_hyperlink(row, 7)
        profile_url  = cell_hyperlink(row, 8)
        youtube_url  = cell_hyperlink(row, 9)
        contact      = clean(cell_val(row, 10))
        email        = clean(cell_val(row, 11))

        recent_lg    = clean(cell_val(row, 12))
        background   = clean(cell_val(row, 13))
        strengths    = clean(cell_val(row, 14))
        lg_exp       = clean(cell_val(row, 15))
        success      = clean(cell_val(row, 16))
        domain       = clean(cell_val(row, 17))

        vis_strat    = clean_bool(cell_val(row, 18))
        perf_exec    = clean_bool(cell_val(row, 19))
        talent       = clean_bool(cell_val(row, 20))
        self_d       = clean_bool(cell_val(row, 21))
        c_level      = clean_bool(cell_val(row, 22))
        exec_coach   = clean_bool(cell_val(row, 23))
        mid_mgr      = clean_bool(cell_val(row, 24))
        diag_tools   = clean(cell_val(row, 25))
        grp_coach    = clean_bool(cell_val(row, 26))
        regional     = clean_bool(cell_val(row, 27))
        languages    = clean(cell_val(row, 28))
        lg_elec      = clean(cell_val(row, 29))
        cert_kca     = clean(cell_val(row, 30))
        cert_other   = clean(cell_val(row, 31))
        cert_icf     = clean(cell_val(row, 32))

        # 가격 매핑
        if   group == 'CiT':       prices = p_cit
        elif group == '코칭경영원': prices = p_jaky
        elif group == '한스코칭':   prices = p_hans.get(name, p_hans.get(name+' 코치', []))
        elif group == '개별':       prices = p_indi.get(name, p_indi.get(name+'(인사)', p_indi.get(name+'(부산)', [])))
        else: prices = []

        coaches.append({
            "no": no, "firm": firm, "group": group, "name": name, "alias": alias,
            "birthYear": by, "gender": gender,
            "hasPhoto": photo_file is not None,
            "profileUrl": profile_url, "introVideoUrl": intro_link, "youtubeUrl": youtube_url,
            "contact": contact, "email": email,
            "recentLgActivity": recent_lg, "background": background,
            "strengths": strengths, "lgExperience": lg_exp,
            "successCases": success, "coachingDomain": domain,
            "visionStrategy": vis_strat, "performanceExec": perf_exec,
            "talentOrg": talent, "selfDev": self_d,
            "cLevel": c_level, "execCoaching": exec_coach, "midManager": mid_mgr,
            "diagnosticTools": diag_tools,
            "groupCoaching": grp_coach, "regionalTravel": regional, "languages": languages,
            "lgElecExperience": lg_elec,
            "certKca": cert_kca, "certDomesticOther": cert_other, "certIcf": cert_icf,
            "prices": prices,
        })

    coaches += sunbae
    coaches.sort(key=lambda c: (c['group'], c['no']))
    print(f"총 {len(coaches)}명")
    for g, cnt in sorted(Counter(c['group'] for c in coaches).items()):
        print(f"  {g}: {cnt}명")

    # 가격 샘플 확인
    sample = next((c for c in coaches if c['prices']), None)
    if sample:
        print(f"\n가격 샘플 ({sample['name']}): {sample['prices'][:2]}")
    else:
        print("\n⚠ 가격 데이터 없음")

    output = f"// Auto-generated by generate_data.py\nconst COACHES = {json.dumps(coaches, ensure_ascii=False, indent=2)};\n"
    with open(OUTPUT_PATH, 'w', encoding='utf-8-sig') as f:
        f.write(output)
    print(f"\ndata.js 생성 완료: {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
